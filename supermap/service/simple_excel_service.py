import os
import logging
import tempfile
from openpyxl import load_workbook
from PIL import Image as PILImage
import io
import boto3
from mimetypes import guess_type
from datetime import datetime
import hashlib
import base64

logger = logging.getLogger(__name__)


class SimpleExcelService:
    """简化的Excel图片读取服务"""
    
    @staticmethod
    def preview_excel_data(excel_file_path):
        """预览Excel文件中的数据"""
        try:
            # 加载Excel文件
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            
            # 获取表头（第一行）
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            logger.info(f"Excel表头: {headers}")
            
            # 定义CheckInPoint模型的字段映射
            field_mapping = {
                'title': 'title',
                'title_en': 'title_en',
                'title_cn': 'title_cn',
                'category': 'category',
                'currency': 'currency',
                'balance': 'balance',
                'level': 'level',
                'hash_rate': 'hash_rate',
                'landmark_image': 'landmark_image',
                'country': 'country',
                'region': 'region',
                'city': 'city',
                'place_id': 'place_id',
                'location': 'location',
                'latitude': 'latitude',
                'longitude': 'longitude',
                'total_checked_in': 'total_checked_in',
                'status': 'status',
                'owner_id': 'owner_id',
                'verification_code': 'verification_code',
                'description': 'description',
                'description_en': 'description_en',
                'description_cn': 'description_cn',
                'is_active': 'is_active',
                'external_url': 'external_url',
                'telegram_url': 'telegram_url',
                'believe_url': 'believe_url',
                'pcn_url': 'pcn_url',
                'whatsapp_url': 'whatsapp_url',
                'audit_status': 'audit_status',
                'believe_id': 'believe_id',
                'wechat': 'wechat'
            }
            
            # 创建列索引映射
            column_mapping = {}
            for col_idx, header in enumerate(headers, 1):
                if header in field_mapping:
                    column_mapping[col_idx] = field_mapping[header]
            
            logger.info(f"列映射: {column_mapping}")
            
            # 处理数据行（从第二行开始）
            preview_data = []
            
            for row_idx in range(2, worksheet.max_row + 1):
                try:
                    # 提取行数据
                    row_data = {}
                    for col_idx, field_name in column_mapping.items():
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        # 过滤掉空值
                        if cell_value is not None and str(cell_value).strip() != '':
                            row_data[field_name] = cell_value
                    
                    # 从R列到AK列提取图片
                    row_images = SimpleExcelService.extract_images_from_row(worksheet, row_idx, excel_file_path)
                    
                    # 添加预览数据
                    preview_data.append({
                        'row_number': row_idx,
                        'data': row_data,
                        'images': row_images,
                        'images_count': len(row_images),
                        'has_images': len(row_images) > 0
                    })
                    
                    logger.info(f"预览第{row_idx}行数据: {row_data.get('title', '无标题')}, 图片数量: {len(row_images)}")
                    
                except Exception as e:
                    logger.error(f"预览第{row_idx}行数据失败: {e}")
                    continue
            
            return {
                'success': True,
                'message': f'成功预览 {len(preview_data)} 行数据',
                'headers': headers,
                'preview_data': preview_data
            }
            
        except Exception as e:
            logger.error(f"预览Excel数据失败: {e}")
            return {
                'success': False,
                'message': f'预览失败: {str(e)}'
            }

    @staticmethod
    def convert_image_to_base64(image_data, max_size=(200, 200)):
        """将图片数据转换为base64格式，用于网页预览"""
        try:
            # 检查图片数据是否有效
            if not image_data or len(image_data) == 0:
                return None
            
            # 使用PIL处理图片
            image = PILImage.open(io.BytesIO(image_data))
            
            # 调整图片大小以适应预览
            image.thumbnail(max_size, PILImage.Resampling.LANCZOS)
            
            # 转换为RGB模式（如果是RGBA，去除透明通道）
            if image.mode in ('RGBA', 'LA', 'P'):
                # 创建白色背景
                background = PILImage.new('RGB', image.size, (255, 255, 255))
                if image.mode == 'P':
                    image = image.convert('RGBA')
                background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
                image = background
            
            # 保存为JPEG格式
            output_buffer = io.BytesIO()
            image.save(output_buffer, format='JPEG', quality=85)
            output_buffer.seek(0)
            
            # 转换为base64
            base64_data = base64.b64encode(output_buffer.getvalue()).decode('utf-8')
            return base64_data
            
        except Exception as e:
            logger.error(f"转换图片为base64失败: {e}")
            return None

    @staticmethod
    def extract_images_from_cell_images(worksheet, row_idx):
        """从单元格内的图片中提取图片数据"""
        row_images = []
        
        try:
            # 获取workbook中的图片
            workbook = worksheet.parent
            if not hasattr(workbook, '_images') or not workbook._images:
                logger.info(f"第{row_idx}行：workbook中没有图片")
                return row_images
            
            logger.info(f"第{row_idx}行：workbook中有 {len(workbook._images)} 张图片")
            
            # 遍历T列到AK列（第20列到第37列）
            for col_idx in range(20, 38):
                col_letter = SimpleExcelService.get_column_letter(col_idx)
                logger.info(f"检查第{row_idx}行{col_letter}列（第{col_idx}列）")
                
                # 检查该单元格是否有图片
                found_image = False
                for image in workbook._images:
                    if hasattr(image, 'ref') and image.ref:
                        # 解析单元格引用
                        import re
                        match = re.match(r'([A-Z]+)(\d+)', image.ref)
                        if match:
                            img_col_letter, img_row_num = match.groups()
                            logger.info(f"  比较：图片ref={image.ref}, 目标={col_letter}{row_idx}")
                            if img_col_letter == col_letter and int(img_row_num) == row_idx:
                                logger.info(f"  找到匹配！图片ref={image.ref}")
                                try:
                                    # 获取图片数据
                                    image_data = image._data()
                                    
                                    if image_data and len(image_data) > 0:
                                        # 生成base64预览
                                        base64_preview = SimpleExcelService.convert_image_to_base64(image_data)
                                        
                                        row_images.append({
                                            'col_letter': col_letter,
                                            'col_num': col_idx,
                                            'size': len(image_data),
                                            'base64_preview': base64_preview,
                                            'has_preview': base64_preview is not None
                                        })
                                        
                                        logger.info(f"第{row_idx}行{col_letter}列：找到图片，大小{len(image_data)}字节")
                                        found_image = True
                                        break  # 找到图片后跳出内层循环
                                    else:
                                        logger.warning(f"第{row_idx}行{col_letter}列：图片数据为空")
                                except Exception as e:
                                    logger.error(f"读取第{row_idx}行{col_letter}列图片数据失败: {e}")
                
                if not found_image:
                    logger.info(f"第{row_idx}行{col_letter}列：没有找到图片")
            
            logger.info(f"第{row_idx}行总共找到 {len(row_images)} 张图片")
            return row_images
            
        except Exception as e:
            logger.error(f"提取第{row_idx}行图片数据失败: {e}")
            return []

    @staticmethod
    def extract_images_from_row(worksheet, row_idx, excel_file_path=None):
        """从指定行提取R列和S列的单元格内图片数据 - 增强版本"""
        row_images = []
        
        try:
            logger.info(f"第{row_idx}行：开始提取R列到AK列的单元格内图片")
            
            # 方法1: 检查DISPIMG函数
            logger.info(f"第{row_idx}行：检查DISPIMG函数")
            dispimg_images = SimpleExcelService.extract_images_from_dispimg(worksheet, row_idx, excel_file_path)
            if dispimg_images:
                row_images.extend(dispimg_images)
                logger.info(f"第{row_idx}行：从DISPIMG函数找到 {len(dispimg_images)} 张图片")
            
            # 方法2: 检查worksheet的drawing
            if not row_images:
                logger.info(f"第{row_idx}行：检查worksheet的drawing")
                drawing_images = SimpleExcelService.extract_images_from_drawing(worksheet, row_idx)
                if drawing_images:
                    row_images.extend(drawing_images)
                    logger.info(f"第{row_idx}行：从drawing找到 {len(drawing_images)} 张图片")
            
            # 方法3: 检查workbook的_images
            if not row_images:
                logger.info(f"第{row_idx}行：检查workbook的_images")
                workbook_images = SimpleExcelService.extract_images_from_workbook(worksheet, row_idx)
                if workbook_images:
                    row_images.extend(workbook_images)
                    logger.info(f"第{row_idx}行：从workbook找到 {len(workbook_images)} 张图片")
            
            logger.info(f"第{row_idx}行总共找到 {len(row_images)} 张图片")
            return row_images
            
        except Exception as e:
            logger.error(f"提取第{row_idx}行图片数据失败: {e}")
            return []

    @staticmethod
    def extract_images_from_dispimg(worksheet, row_idx, excel_file_path=None):
        """从DISPIMG函数中提取图片数据"""
        row_images = []
        
        try:
            # 检查T列到AK列（第20列到第37列）
            target_columns = []
            for col_idx in range(20, 38):  # T列(20)到AK列(37)
                col_letter = SimpleExcelService.get_column_letter(col_idx)
                target_columns.append((col_idx, col_letter))
            
            for col_idx, col_letter in target_columns:
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                
                # 如果单元格为空，跳过
                if not cell_value or str(cell_value).strip() == '':
                    logger.info(f"第{row_idx}行{col_letter}列：单元格为空，跳过")
                    continue
                
                if isinstance(cell_value, str) and 'DISPIMG' in cell_value:
                    # 解析DISPIMG函数获取图片ID
                    import re
                    dispimg_match = re.search(r'DISPIMG\("([^"]+)"', cell_value)
                    if dispimg_match:
                        image_id = dispimg_match.group(1)
                        logger.info(f"第{row_idx}行{col_letter}列：找到DISPIMG函数，图片ID: {image_id}")
                        
                        # 尝试从Excel文件中提取图片数据
                        # 根据DISPIMG函数中的image_id来匹配cellImage
                        image_data = SimpleExcelService.extract_image_from_excel_file_by_image_id(worksheet, image_id, excel_file_path)
                        
                        if image_data:
                            # 生成base64预览
                            base64_preview = SimpleExcelService.convert_image_to_base64(image_data)
                            
                            row_images.append({
                                'col_letter': col_letter,
                                'col_num': col_idx,
                                'size': len(image_data),
                                'data': image_data,  # 添加原始图片数据
                                'base64_preview': base64_preview,
                                'has_preview': base64_preview is not None,
                                'type': 'dispimg',
                                'image_id': image_id
                            })
                            
                            logger.info(f"第{row_idx}行{col_letter}列：成功提取图片，大小{len(image_data)}字节")
                        else:
                            logger.warning(f"第{row_idx}行{col_letter}列：无法提取图片数据")
            
            return row_images
            
        except Exception as e:
            logger.error(f"从DISPIMG函数提取图片失败: {e}")
            return []

    @staticmethod
    def extract_image_from_excel_file_by_image_id(worksheet, image_id, excel_file_path=None):
        """根据DISPIMG函数中的image_id从Excel文件中提取图片数据"""
        try:
            # 获取workbook
            workbook = worksheet.parent
            
            # 方法1: 尝试从workbook的_images中获取
            if hasattr(workbook, '_images') and workbook._images:
                for image in workbook._images:
                    if hasattr(image, 'id') and image.id == image_id:
                        try:
                            image_data = image._data()
                            if image_data and len(image_data) > 0:
                                logger.info(f"从workbook._images找到图片ID {image_id}")
                                return image_data
                        except Exception as e:
                            logger.error(f"读取workbook图片数据失败: {e}")
            
            # 方法2: 尝试从Excel文件内部结构中获取
            if excel_file_path and os.path.exists(excel_file_path):
                try:
                    import zipfile
                    logger.info(f"尝试从Excel文件内部结构提取图片ID {image_id}: {excel_file_path}")
                    
                    with zipfile.ZipFile(excel_file_path, 'r') as zip_file:
                        # 读取cellimages.xml
                        if 'xl/cellimages.xml' in zip_file.namelist():
                            with zip_file.open('xl/cellimages.xml') as xml_file:
                                import xml.etree.ElementTree as ET
                                root = ET.fromstring(xml_file.read())
                                
                                # 查找所有cellImage元素 - 包括WPS命名空间
                                cell_images = (
                                    root.findall('.//cellImage') + 
                                    root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cellImage') +
                                    root.findall('.//{http://www.wps.cn/officeDocument/2017/etCustomData}cellImage')
                                )
                                logger.info(f"找到 {len(cell_images)} 个cellImage元素")
                                
                                # 根据image_id匹配对应的cellImage
                                for i, cell_image in enumerate(cell_images):
                                    logger.info(f"检查cellImage {i+1}")
                                    
                                    # 从cellImage中提取name属性（对应DISPIMG中的image_id）
                                    cell_image_name = None
                                    for child in cell_image:
                                        if 'pic' in child.tag:
                                            for pic_child in child:
                                                if 'nvPicPr' in pic_child.tag:
                                                    for nv_child in pic_child:
                                                        if 'cNvPr' in nv_child.tag:
                                                            cell_image_name = nv_child.get('name')
                                                            if cell_image_name:
                                                                logger.info(f"  cellImage {i+1} 的name: {cell_image_name}")
                                                                break
                                                    if cell_image_name:
                                                        break
                                            if cell_image_name:
                                                break
                                    
                                    # 如果找到匹配的image_id
                                    if cell_image_name == image_id:
                                        logger.info(f"找到匹配的cellImage {i+1}，image_id: {image_id}")
                                        
                                        # 从cellImage中提取embed_id
                                        img_id = None
                                        for child in cell_image:
                                            if 'pic' in child.tag:
                                                for pic_child in child:
                                                    if 'blipFill' in pic_child.tag:
                                                        for blip_child in pic_child:
                                                            if 'blip' in blip_child.tag:
                                                                embed = blip_child.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                                                if embed:
                                                                    img_id = embed
                                                                    logger.info(f"找到embed ID: {img_id}")
                                                                    
                                                                    # 使用embed ID获取图片数据
                                                                    image_data = SimpleExcelService.get_image_data_by_embed_id(worksheet, img_id, excel_file_path)
                                                                    if image_data:
                                                                        logger.info(f"成功获取图片ID {image_id} 的图片数据")
                                                                        return image_data
                                                                    break
                                                        if img_id:
                                                            break
                                                if img_id:
                                                    break
                                        break
                                
                except Exception as e:
                    logger.error(f"从Excel文件内部结构提取图片失败: {e}")
            
            logger.warning(f"无法找到图片ID {image_id} 的图片")
            return None
            
        except Exception as e:
            logger.error(f"从Excel文件提取图片失败: {e}")
            return None

    @staticmethod
    def extract_image_from_excel_file_by_position(worksheet, col_idx, excel_file_path=None):
        """根据列位置从Excel文件中提取图片数据"""
        try:
            # 获取workbook
            workbook = worksheet.parent
            
            # 方法1: 尝试从workbook的_images中获取
            if hasattr(workbook, '_images') and workbook._images:
                for image in workbook._images:
                    if hasattr(image, 'ref') and image.ref:
                        import re
                        match = re.match(r'([A-Z]+)(\d+)', image.ref)
                        if match:
                            img_col_letter, img_row_num = match.groups()
                            col_letter = SimpleExcelService.get_column_letter(col_idx)
                            if img_col_letter == col_letter:
                                try:
                                    image_data = image._data()
                                    if image_data and len(image_data) > 0:
                                        logger.info(f"从workbook._images找到列{col_letter}的图片")
                                        return image_data
                                except Exception as e:
                                    logger.error(f"读取workbook图片数据失败: {e}")
            
            # 方法2: 尝试从Excel文件内部结构中获取
            if excel_file_path and os.path.exists(excel_file_path):
                try:
                    import zipfile
                    logger.info(f"尝试从Excel文件内部结构提取列{col_idx}的图片: {excel_file_path}")
                    
                    with zipfile.ZipFile(excel_file_path, 'r') as zip_file:
                        # 读取cellimages.xml
                        if 'xl/cellimages.xml' in zip_file.namelist():
                            with zip_file.open('xl/cellimages.xml') as xml_file:
                                import xml.etree.ElementTree as ET
                                root = ET.fromstring(xml_file.read())
                                
                                # 查找所有cellImage元素 - 包括WPS命名空间
                                cell_images = (
                                    root.findall('.//cellImage') + 
                                    root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cellImage') +
                                    root.findall('.//{http://www.wps.cn/officeDocument/2017/etCustomData}cellImage')
                                )
                                logger.info(f"找到 {len(cell_images)} 个cellImage元素")
                                
                                # 根据列位置选择对应的cellImage
                                # T列(20)对应第一个cellImage，U列(21)对应第二个，以此类推
                                cell_image_index = col_idx - 20  # T列是第20列，所以减去20
                                
                                if 0 <= cell_image_index < len(cell_images):
                                    cell_image = cell_images[cell_image_index]
                                    logger.info(f"选择cellImage {cell_image_index + 1} 对应列{col_idx}")
                                    
                                    # 从cellImage中提取embed_id
                                    img_id = None
                                    for child in cell_image:
                                        if 'pic' in child.tag:
                                            for pic_child in child:
                                                if 'blipFill' in pic_child.tag:
                                                    for blip_child in pic_child:
                                                        if 'blip' in blip_child.tag:
                                                            embed = blip_child.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                                            if embed:
                                                                img_id = embed
                                                                logger.info(f"找到embed ID: {img_id}")
                                                                
                                                                # 使用embed ID获取图片数据
                                                                image_data = SimpleExcelService.get_image_data_by_embed_id(worksheet, img_id, excel_file_path)
                                                                if image_data:
                                                                    logger.info(f"成功获取列{col_idx}的图片数据")
                                                                    return image_data
                                                                break
                                                    if img_id:
                                                        break
                                            if img_id:
                                                break
                                else:
                                    logger.warning(f"列{col_idx}超出cellImage范围")
                                
                except Exception as e:
                    logger.error(f"从Excel文件内部结构提取图片失败: {e}")
            
            logger.warning(f"无法找到列{col_idx}的图片")
            return None
            
        except Exception as e:
            logger.error(f"从Excel文件提取图片失败: {e}")
            return None

    @staticmethod
    def extract_image_from_excel_file(worksheet, image_id, excel_file_path=None):
        """从Excel文件中提取指定ID的图片数据"""
        try:
            # 获取workbook
            workbook = worksheet.parent
            
            # 方法1: 尝试从workbook的_images中获取
            if hasattr(workbook, '_images') and workbook._images:
                for image in workbook._images:
                    if hasattr(image, 'id') and image.id == image_id:
                        try:
                            image_data = image._data()
                            if image_data and len(image_data) > 0:
                                logger.info(f"从workbook._images找到图片ID {image_id}")
                                return image_data
                        except Exception as e:
                            logger.error(f"读取workbook图片数据失败: {e}")
            
            # 方法2: 尝试从Excel文件内部结构中获取
            if excel_file_path and os.path.exists(excel_file_path):
                try:
                    import zipfile
                    logger.info(f"尝试从Excel文件内部结构提取图片: {excel_file_path}")
                    
                    with zipfile.ZipFile(excel_file_path, 'r') as zip_file:
                        # 读取cellimages.xml
                        if 'xl/cellimages.xml' in zip_file.namelist():
                            with zip_file.open('xl/cellimages.xml') as xml_file:
                                import xml.etree.ElementTree as ET
                                root = ET.fromstring(xml_file.read())
                                
                                # 查找匹配的图片ID - 使用更宽松的查找方式
                                logger.info(f"在cellimages.xml中查找图片ID: {image_id}")
                                
                                # 查找所有cellImage元素 - 包括WPS命名空间
                                cell_images = (
                                    root.findall('.//cellImage') + 
                                    root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}cellImage') +
                                    root.findall('.//{http://www.wps.cn/officeDocument/2017/etCustomData}cellImage')
                                )
                                logger.info(f"找到 {len(cell_images)} 个cellImage元素")
                                
                                # 如果还是找不到，尝试查找所有可能的元素
                                if len(cell_images) == 0:
                                    logger.info("尝试查找所有可能的图片相关元素")
                                    all_elements = root.findall('.//*')
                                    for elem in all_elements:
                                        if 'image' in elem.tag.lower() or 'cell' in elem.tag.lower():
                                            logger.info(f"找到相关元素: {elem.tag}")
                                            if elem.get('imageId'):
                                                logger.info(f"  有imageId: {elem.get('imageId')}")
                                            if elem.get('id'):
                                                logger.info(f"  有id: {elem.get('id')}")
                                
                                for cell_image in cell_images:
                                    # 检查所有可能的属性
                                    logger.info(f"检查cellImage: {cell_image.tag}")
                                    logger.info(f"  所有属性: {cell_image.attrib}")
                                    
                                    # 初始化img_id
                                    img_id = None
                                    
                                    # 检查子元素
                                    logger.info(f"  子元素数量: {len(cell_image)}")
                                    for i, child in enumerate(cell_image):
                                        logger.info(f"  子元素 {i}: {child.tag}, 属性: {child.attrib}")
                                        
                                        # 检查pic元素的子元素
                                        if 'pic' in child.tag:
                                            logger.info(f"    检查pic元素的子元素:")
                                            for j, pic_child in enumerate(child):
                                                logger.info(f"      pic子元素 {j}: {pic_child.tag}, 属性: {pic_child.attrib}")
                                                
                                                # 检查blipFill元素的子元素
                                                if 'blipFill' in pic_child.tag:
                                                    logger.info(f"        检查blipFill的子元素:")
                                                    for k, blip_child in enumerate(pic_child):
                                                        logger.info(f"          blipFill子元素 {k}: {blip_child.tag}, 属性: {blip_child.attrib}")
                                                        
                                                        # 检查是否有blip元素
                                                        if 'blip' in blip_child.tag:
                                                            logger.info(f"            找到blip元素: {blip_child.attrib}")
                                                            # 检查embed属性
                                                            embed = blip_child.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                                            if embed:
                                                                logger.info(f"            找到embed: {embed}")
                                                                # 这里应该能找到图片ID
                                                                img_id = embed
                                    
                                    # 尝试不同的属性名
                                    if not img_id:
                                        img_id = cell_image.get('imageId') or cell_image.get('id') or cell_image.get('image_id')
                                    logger.info(f"检查cellImage: imageId={img_id}")
                                    # 如果找到了embed ID，直接使用它来获取图片数据
                                    if img_id and img_id.startswith('rId'):
                                        logger.info(f"找到embed ID: {img_id}，尝试获取图片数据")
                                        # 直接使用embed ID获取图片数据
                                        image_data = SimpleExcelService.get_image_data_by_embed_id(worksheet, img_id, excel_file_path)
                                        if image_data:
                                            return image_data
                                    
                                    # 原来的逻辑：检查是否匹配DISPIMG中的图片ID
                                    if img_id == image_id:
                                        logger.info(f"找到匹配的cellImage: {img_id}")
                                        # 获取图片文件路径
                                        images = root.findall('.//image') + root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}image')
                                        logger.info(f"找到 {len(images)} 个image元素")
                                        
                                        for image in images:
                                            img_image_id = image.get('id')
                                            logger.info(f"检查image: id={img_image_id}")
                                            if img_image_id == img_id:
                                                for child in image:
                                                    if child.tag.endswith('}blip'):
                                                        embed = child.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                                        if embed:
                                                            # 读取关系文件
                                                            if 'xl/_rels/cellimages.xml.rels' in zip_file.namelist():
                                                                with zip_file.open('xl/_rels/cellimages.xml.rels') as rels_file:
                                                                    rels_root = ET.fromstring(rels_file.read())
                                                                    for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                                                        if rel.get('Id') == embed:
                                                                            target = rel.get('Target')
                                                                            if target:
                                                                                # 读取图片文件
                                                                                image_path = f"xl/{target}"
                                                                                if image_path in zip_file.namelist():
                                                                                    with zip_file.open(image_path) as img_file:
                                                                                        image_data = img_file.read()
                                                                                        logger.info(f"从Excel文件内部结构找到图片ID {image_id}")
                                                                                        return image_data
                except Exception as e:
                    logger.error(f"从Excel文件内部结构提取图片失败: {e}")
            
            logger.warning(f"无法找到图片ID {image_id}")
            return None
            
        except Exception as e:
            logger.error(f"从Excel文件提取图片失败: {e}")
            return None

    @staticmethod
    def get_image_data_by_embed_id(worksheet, embed_id, excel_file_path):
        """通过embed ID获取图片数据"""
        try:
            logger.info(f"通过embed ID {embed_id} 获取图片数据")
            
            # 从关系文件中获取图片路径
            import zipfile
            with zipfile.ZipFile(excel_file_path, 'r') as zip_file:
                # 读取关系文件
                if 'xl/_rels/cellimages.xml.rels' in zip_file.namelist():
                    with zip_file.open('xl/_rels/cellimages.xml.rels') as rels_file:
                        import xml.etree.ElementTree as ET
                        rels_root = ET.fromstring(rels_file.read())
                        
                        # 查找匹配的关系
                        for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                            if rel.get('Id') == embed_id:
                                target = rel.get('Target')
                                if target:
                                    logger.info(f"找到关系 {embed_id} -> {target}")
                                    
                                    # 读取图片文件
                                    image_path = f"xl/{target}"
                                    if image_path in zip_file.namelist():
                                        with zip_file.open(image_path) as img_file:
                                            image_data = img_file.read()
                                            logger.info(f"成功读取图片数据，大小: {len(image_data)} 字节")
                                            return image_data
                                    else:
                                        logger.warning(f"图片文件不存在: {image_path}")
                                else:
                                    logger.warning(f"关系 {embed_id} 没有target属性")
                        else:
                            logger.warning(f"找不到关系ID: {embed_id}")
                else:
                    logger.warning("找不到关系文件: xl/_rels/cellimages.xml.rels")
            
            return None
            
        except Exception as e:
            logger.error(f"通过embed ID获取图片数据失败: {e}")
            return None

    @staticmethod
    def extract_images_from_drawing(worksheet, row_idx):
        """从worksheet的drawing中提取图片数据"""
        row_images = []
        
        try:
            # 检查worksheet是否有drawing
            if not hasattr(worksheet, 'drawing') or not worksheet.drawing:
                return row_images
            
            # 检查drawing中是否有图片
            if not hasattr(worksheet.drawing, 'images') or not worksheet.drawing.images:
                return row_images
            
            logger.info(f"第{row_idx}行：drawing中有 {len(worksheet.drawing.images)} 张图片")
            
            # 获取workbook以访问关系
            workbook = worksheet.parent
            
            # 专门检查T列和U列（第20列和第21列）
            target_columns = [(20, 'T'), (21, 'U')]
            
            for col_idx, col_letter in target_columns:
                # 遍历drawing中的所有图片
                for i, image in enumerate(worksheet.drawing.images):
                    try:
                        # 获取图片的位置信息
                        if hasattr(image, 'anchor'):
                            anchor = image.anchor
                            
                            # 获取图片的单元格位置
                            if hasattr(anchor, '_from'):
                                from_pos = anchor._from
                                
                                # 检查是否在目标行和列
                                if hasattr(from_pos, 'row') and from_pos.row == row_idx - 1:  # openpyxl行号从0开始
                                    img_col_idx = from_pos.col + 1  # openpyxl列号从0开始
                                    
                                    if img_col_idx == col_idx:
                                        logger.info(f"第{row_idx}行{col_letter}列：找到drawing图片")
                                        
                                        # 通过关系ID获取图片数据
                                        if hasattr(image, 'ref'):
                                            relationship_id = image.ref
                                            
                                            # 从workbook的关系中获取图片数据
                                            if hasattr(workbook, '_rels') and workbook._rels:
                                                for rel_id, rel in workbook._rels.items():
                                                    if rel_id == relationship_id:
                                                        # 获取图片文件路径
                                                        if hasattr(rel, 'target'):
                                                            image_path = rel.target
                                                            
                                                            # 从workbook中读取图片文件
                                                            try:
                                                                # 获取图片数据
                                                                image_data = workbook._images[relationship_id]._data()
                                                                
                                                                if image_data and len(image_data) > 0:
                                                                    # 生成base64预览
                                                                    base64_preview = SimpleExcelService.convert_image_to_base64(image_data)
                                                                    
                                                                    row_images.append({
                                                                        'col_letter': col_letter,
                                                                        'col_num': col_idx,
                                                                        'size': len(image_data),
                                                                        'base64_preview': base64_preview,
                                                                        'has_preview': base64_preview is not None,
                                                                        'relationship_id': relationship_id,
                                                                        'image_path': image_path,
                                                                        'type': 'drawing'
                                                                    })
                                                                    
                                                                    logger.info(f"第{row_idx}行{col_letter}列：找到drawing图片，大小{len(image_data)}字节")
                                                                    break  # 找到图片后跳出循环
                                                            except Exception as e:
                                                                logger.error(f"读取drawing图片数据失败: {e}")
                            
                    except Exception as e:
                        logger.error(f"处理drawing图片 {i} 失败: {e}")
            
            return row_images
            
        except Exception as e:
            logger.error(f"从drawing提取图片失败: {e}")
            return []

    @staticmethod
    def extract_images_from_workbook(worksheet, row_idx):
        """从workbook的_images中提取图片数据"""
        row_images = []
        
        try:
            # 获取workbook中的图片
            workbook = worksheet.parent
            if not hasattr(workbook, '_images') or not workbook._images:
                return row_images
            
            logger.info(f"第{row_idx}行：workbook中有 {len(workbook._images)} 张图片")
            
            # 遍历T列到U列（第20列到第21列）
            for col_idx in range(20, 21):
                col_letter = SimpleExcelService.get_column_letter(col_idx)
                
                # 检查该单元格是否有图片
                for image in workbook._images:
                    if hasattr(image, 'ref') and image.ref:
                        # 解析单元格引用
                        import re
                        match = re.match(r'([A-Z]+)(\d+)', image.ref)
                        if match:
                            img_col_letter, img_row_num = match.groups()
                            if img_col_letter == col_letter and int(img_row_num) == row_idx:
                                try:
                                    # 获取图片数据
                                    image_data = image._data()
                                    
                                    if image_data and len(image_data) > 0:
                                        # 生成base64预览
                                        base64_preview = SimpleExcelService.convert_image_to_base64(image_data)
                                        
                                        row_images.append({
                                            'col_letter': col_letter,
                                            'col_num': col_idx,
                                            'size': len(image_data),
                                            'base64_preview': base64_preview,
                                            'has_preview': base64_preview is not None,
                                            'type': 'workbook'
                                        })
                                        
                                        logger.info(f"第{row_idx}行{col_letter}列：找到workbook图片，大小{len(image_data)}字节")
                                        break  # 找到图片后跳出循环
                                except Exception as e:
                                    logger.error(f"读取workbook图片数据失败: {e}")
            
            return row_images
            
        except Exception as e:
            logger.error(f"从workbook提取图片失败: {e}")
            return []

    @staticmethod
    def debug_excel_images(excel_file_path, target_row=2):
        """调试Excel文件中R2和S2单元格的图片结构"""
        try:
            logger.info(f"=== 开始调试Excel文件: {excel_file_path} ===")
            
            # 加载Excel文件
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            
            logger.info(f"工作表名称: {worksheet.title}")
            logger.info(f"最大行数: {worksheet.max_row}")
            logger.info(f"最大列数: {worksheet.max_column}")
            
            # 检查T2和U2单元格的内容
            logger.info(f"检查第{target_row}行R和S列的单元格内容:")
            for col_idx, col_letter in [(20, 'T'), (21, 'U')]:
                cell_value = worksheet.cell(row=target_row, column=col_idx).value
                logger.info(f"  {col_letter}列: {cell_value}")
            
            # 检查workbook的关系
            if hasattr(workbook, '_rels') and workbook._rels:
                logger.info(f"workbook中有 {len(workbook._rels)} 个关系")
                for rel_id, rel in workbook._rels.items():
                    logger.info(f"关系 {rel_id}:")
                    logger.info(f"  - 类型: {type(rel)}")
                    if hasattr(rel, 'target'):
                        logger.info(f"  - 目标: {rel.target}")
                    if hasattr(rel, 'reltype'):
                        logger.info(f"  - 关系类型: {rel.reltype}")
                    if hasattr(rel, 'rId'):
                        logger.info(f"  - 关系ID: {rel.rId}")
            else:
                logger.info("workbook中没有关系")
            
            # 检查workbook中的图片
            if hasattr(workbook, '_images') and workbook._images:
                logger.info(f"workbook中有 {len(workbook._images)} 张图片")
                for i, image in enumerate(workbook._images):
                    logger.info(f"图片 {i}:")
                    logger.info(f"  - 类型: {type(image)}")
                    if hasattr(image, 'ref'):
                        logger.info(f"  - ref: {image.ref}")
                    if hasattr(image, 'id'):
                        logger.info(f"  - id: {image.id}")
                    if hasattr(image, 'name'):
                        logger.info(f"  - name: {image.name}")
                    try:
                        if hasattr(image, '_data'):
                            data = image._data()
                            logger.info(f"  - 数据大小: {len(data)} 字节")
                            logger.info(f"  - 数据前20字节: {data[:20]}")
                    except Exception as e:
                        logger.error(f"  - 读取数据失败: {e}")
            else:
                logger.info("workbook中没有图片")
            
            # 检查worksheet的drawing
            if hasattr(worksheet, 'drawing') and worksheet.drawing:
                logger.info("worksheet有drawing属性")
                if hasattr(worksheet.drawing, 'images') and worksheet.drawing.images:
                    logger.info(f"drawing中有 {len(worksheet.drawing.images)} 张图片")
                    for i, image in enumerate(worksheet.drawing.images):
                        logger.info(f"drawing图片 {i}:")
                        logger.info(f"  - 类型: {type(image)}")
                        if hasattr(image, 'ref'):
                            logger.info(f"  - ref: {image.ref}")
                        if hasattr(image, 'anchor'):
                            anchor = image.anchor
                            logger.info(f"  - anchor: {anchor}")
                            if hasattr(anchor, '_from'):
                                from_pos = anchor._from
                                logger.info(f"  - 位置: row={from_pos.row}, col={from_pos.col}")
                        try:
                            if hasattr(image, '_data'):
                                data = image._data()
                                logger.info(f"  - 数据大小: {len(data)} 字节")
                        except Exception as e:
                            logger.error(f"  - 读取数据失败: {e}")
                else:
                    logger.info("drawing中没有图片")
            else:
                logger.info("worksheet没有drawing属性")
            
            # 测试图片提取
            logger.info("=== 测试图片提取 ===")
            extracted_images = SimpleExcelService.extract_images_from_row(worksheet, target_row, excel_file_path)
            logger.info(f"提取到的图片数量: {len(extracted_images)}")
            for i, img in enumerate(extracted_images):
                logger.info(f"提取的图片 {i}:")
                logger.info(f"  - 列: {img.get('col_letter')}")
                logger.info(f"  - 大小: {img.get('size')} 字节")
                logger.info(f"  - 有base64预览: {img.get('has_preview')}")
                logger.info(f"  - 关系ID: {img.get('relationship_id')}")
                logger.info(f"  - 图片路径: {img.get('image_path')}")
            
            logger.info(f"=== 调试Excel文件结束 ===")
            
            return {
                'success': True,
                'message': f'成功调试Excel文件，检查了第{target_row}行R和S列',
                'extracted_images_count': len(extracted_images)
            }
            
        except Exception as e:
            logger.error(f"调试Excel文件失败: {e}")
            return {
                'success': False,
                'message': f'调试失败: {str(e)}'
            }

    @staticmethod
    def get_column_letter(col_idx):
        """将列索引转换为列字母"""
        if col_idx <= 26:
            return chr(ord('A') + col_idx - 1)
        else:
            return chr(ord('A') + (col_idx - 1) // 26 - 1) + chr(ord('A') + (col_idx - 1) % 26)

    def __init__(self):
        """初始化S3客户端"""
        self.s3_client = boto3.client(
            's3',
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
            region_name=os.getenv('AWS_S3_REGION_NAME')
        )
        self.bucket_name = os.getenv('AWS_STORAGE_BUCKET_NAME')
        self.domain = os.getenv('AWS_S3_CUSTOM_DOMAIN')

    def get_md5_hash(self, input_str: str) -> str:
        """生成字符串的MD5哈希值"""
        md5_hash = hashlib.md5()
        md5_hash.update(input_str.encode('utf-8'))
        return md5_hash.hexdigest()

    def upload_image_to_s3(self, image_data: bytes, filename: str, category: str = 'excel_import') -> str:
        """上传图片到S3"""
        try:
            # 生成文件名
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            file_ext = filename.split('.')[-1] if '.' in filename else 'jpg'
            # 使用更简洁的文件名格式：时间戳_随机数.扩展名
            import random
            random_num = random.randint(10000, 99999)
            final_filename = f"{category}/{timestamp}_{random_num}.{file_ext}"
            
            # 确定内容类型
            content_type, _ = guess_type(final_filename)
            if not content_type:
                content_type = 'image/jpeg'
            
            # 上传到S3
            self.s3_client.put_object(
                Bucket=self.bucket_name,
                Key=final_filename,
                Body=image_data,
                ContentType=content_type
            )
            
            # 返回完整的URL
            return f"{self.domain}/{final_filename}"
            
        except Exception as e:
            logger.error(f"上传图片到S3失败: {str(e)}")
            raise

    @staticmethod
    def import_excel_data(excel_file_path):
        """导入Excel数据到check_in_point表"""
        from ..models import CheckInPoint, CheckInPointImage
        
        try:
            # 加载Excel文件
            workbook = load_workbook(excel_file_path)
            worksheet = workbook.active
            
            # 获取表头（第一行）
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(str(cell_value).strip())
            
            logger.info(f"Excel表头: {headers}")
            
            # 定义CheckInPoint模型的字段映射
            field_mapping = {
                'title': 'title',
                'title_en': 'title_en',
                'title_cn': 'title_cn',
                'category': 'category',
                'currency': 'currency',
                'balance': 'balance',
                'level': 'level',
                'hash_rate': 'hash_rate',
                'landmark_image': 'landmark_image',
                'country': 'country',
                'region': 'region',
                'city': 'city',
                'place_id': 'place_id',
                'location': 'location',
                'latitude': 'latitude',
                'longitude': 'longitude',
                'total_checked_in': 'total_checked_in',
                'status': 'status',
                'owner_id': 'owner_id',
                'verification_code': 'verification_code',
                'description': 'description',
                'description_en': 'description_en',
                'description_cn': 'description_cn',
                'is_active': 'is_active',
                'external_url': 'external_url',
                'telegram_url': 'telegram_url',
                'believe_url': 'believe_url',
                'pcn_url': 'pcn_url',
                'whatsapp_url': 'whatsapp_url',
                'audit_status': 'audit_status',
                'believe_id': 'believe_id',
                'wechat': 'wechat'
            }
            
            # 创建列索引映射
            column_mapping = {}
            for col_idx, header in enumerate(headers, 1):
                if header in field_mapping:
                    column_mapping[col_idx] = field_mapping[header]
            
            logger.info(f"列映射: {column_mapping}")
            
            # 处理数据行（从第二行开始）
            imported_records = []
            excel_service = SimpleExcelService()
            
            for row_idx in range(2, worksheet.max_row + 1):
                try:
                    # 提取行数据
                    row_data = {}
                    for col_idx, field_name in column_mapping.items():
                        cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                        # 过滤掉空值
                        if cell_value is not None and str(cell_value).strip() != '':
                            row_data[field_name] = cell_value
                    
                    # 处理数据类型转换
                    numeric_fields = ['balance', 'level', 'hash_rate', 'total_checked_in', 'status', 'owner_id', 'is_active', 'audit_status']
                    for field in numeric_fields:
                        if field in row_data and row_data[field] is not None:
                            try:
                                if field in ['level', 'total_checked_in', 'status', 'owner_id', 'is_active', 'audit_status']:
                                    row_data[field] = int(row_data[field])
                                else:
                                    row_data[field] = float(row_data[field])
                            except (ValueError, TypeError):
                                row_data[field] = 0 if field in ['level', 'total_checked_in', 'status', 'owner_id', 'is_active', 'audit_status'] else 0.0
                    
                    # 处理经纬度
                    for field in ['latitude', 'longitude']:
                        if field in row_data and row_data[field] is not None:
                            try:
                                row_data[field] = float(row_data[field])
                            except (ValueError, TypeError):
                                row_data[field] = 0.0
                    
                    # 创建CheckInPoint记录
                    row_data['title'] = row_data['title_en'] if row_data['title_en'] else row_data['title_cn']
                    row_data['description'] = row_data['description_en'] if row_data['description_en'] else row_data['description_cn']
                    row_data['currency'] = 'platform'
                    row_data['balance'] = 0
                    row_data['level'] = 1
                    row_data['hash_rate'] = 0
                    row_data['status'] = 1
                    row_data['is_active'] = 1
                    row_data['audit_status'] = 1
                    check_in_point = CheckInPoint.objects.create(**row_data)
                    
                    # 提取图片并上传
                    row_images = SimpleExcelService.extract_images_from_row(worksheet, row_idx, excel_file_path)
                    
                    # 上传图片到S3并创建CheckInPointImage记录
                    for i, img_info in enumerate(row_images):
                        try:
                            # 获取图片数据
                            if 'data' in img_info and img_info['data']:
                                # 直接使用提取的图片数据
                                image_data = img_info['data']
                            else:
                                # 如果没有data字段，尝试从DISPIMG函数重新提取
                                logger.info(f"尝试重新提取图片数据: 行{row_idx}, 列{img_info['col_letter']}")
                                # 这里可以添加重新提取的逻辑，但通常extract_images_from_row已经处理了
                                continue
                            
                            # 上传到S3
                            image_url = excel_service.upload_image_to_s3(
                                image_data,
                                f"image_{row_idx}_{i}_{img_info['col_letter']}.jpg",
                                f"{check_in_point.category}/{check_in_point.id}"
                            )
                            
                            # 创建CheckInPointImage记录
                            CheckInPointImage.objects.create(
                                check_in_point=check_in_point,
                                image_url=image_url,
                                sort_index=i,
                                status=1
                            )
                            
                            logger.info(f"成功上传图片: {image_url}")
                            
                        except Exception as e:
                            logger.error(f"处理图片失败: {e}")
                            continue
                    
                    imported_records.append({
                        'id': check_in_point.id,
                        'title': check_in_point.title,
                        'images_count': len(row_images)
                    })
                    
                    logger.info(f"成功导入第{row_idx}行数据: {check_in_point.title}")
                    
                except Exception as e:
                    logger.error(f"导入第{row_idx}行数据失败: {e}")
                    continue
            
            return {
                'success': True,
                'message': f'成功导入 {len(imported_records)} 条记录',
                'records': imported_records
            }
            
        except Exception as e:
            logger.error(f"导入Excel数据失败: {e}")
            return {
                'success': False,
                'message': f'导入失败: {str(e)}'
            } 